"""MASAGI Store — digital products landing + purchase pipeline.

Flow (mirrors a Stripe/DOKU Checkout pipeline, self-hosted):
  visitor -> /store -> email capture -> POST /api/store/checkout (order: pending)
          -> /store/pay/<order>  (DOKU checkout page; sandbox simulator included)
          -> gateway notifies POST /api/store/webhook/doku  (HMAC-verified;
             THE ONLY PLACE ACCESS IS GRANTED — never client-side)
          -> customer account auto-created, magic-link + receipt "emails" sent
          -> /store/thanks?order=... polls status -> /downloads dashboard

DOKU integration points are marked with  # DOKU:  comments — set the env vars
DOKU_CLIENT_ID / DOKU_SECRET_KEY and swap the sandbox simulator for the real
Checkout API (https://api-sandbox.doku.com/checkout/v1/payment).

Transactional email: written to data/store_outbox/*.eml.txt by _send_email().
# EMAIL: swap _send_email's body for Resend/Postmark/SMTP — one function.

Files are stored PRIVATELY in data/store_files/ (not under /static, never
listable); customers only ever receive signed, time-limited, use-limited
download tokens.
"""
import base64
import hashlib
import hmac
import io
import json
import os
import secrets
import sqlite3
import time
import zipfile
from datetime import datetime, timedelta

from flask import (Blueprint, jsonify, redirect, request, send_file,
                   send_from_directory, session)
from werkzeug.security import check_password_hash, generate_password_hash

import database

bp = Blueprint("store", __name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
# Lives in its own subfolder so the ERP's multi-database scanner (which treats
# every *.db directly under data/ as an ERP database) never touches it.
STORE_DIR = os.path.join(database.DATA_DIR, "store")
STORE_DB = os.path.join(STORE_DIR, "store.db")
FILES_DIR = os.path.join(STORE_DIR, "files")    # private storage — never public
OUTBOX_DIR = os.path.join(STORE_DIR, "outbox")  # mock email outbox

# DOKU: sandbox credentials — replace via environment for production.
DOKU_CLIENT_ID = os.environ.get("DOKU_CLIENT_ID", "MASAGI-SANDBOX")
DOKU_SECRET_KEY = os.environ.get("DOKU_SECRET_KEY", "sandbox-secret-change-me")

DOWNLOAD_TOKEN_HOURS = 24   # signed links expire after this
DOWNLOAD_TOKEN_MAX_USES = 5  # ... or after this many downloads
MEMBERSHIP_DAYS = 31

# ---------------------------------------------------------------------------
# Catalog — id, name, tagline, price (IDR), type, private file
# ---------------------------------------------------------------------------

PRODUCTS = {
    "coa-template": {
        "name": "SME Chart of Accounts Template", "kind": "Excel template",
        "tagline": "The standard COA we deploy at every MASAGI HV client — ready to import.",
        "price": 149_000, "file": "masagi-coa-template.xlsx", "icon": "≡",
    },
    "budget-planner": {
        "name": "Budget & Cashflow Planner", "kind": "Excel template",
        "tagline": "12-month budget grid with realization tracking, the MASAGI way.",
        "price": 179_000, "file": "masagi-budget-planner.xlsx", "icon": "◧",
    },
    "finance-playbook": {
        "name": "SME Finance Playbook", "kind": "Ebook · PDF",
        "tagline": "How growing Indonesian SMEs get their numbers under control.",
        "price": 129_000, "file": "masagi-sme-finance-playbook.pdf", "icon": "▤",
    },
    "close-checklist": {
        "name": "Month-End Close Checklist", "kind": "PDF",
        "tagline": "Close the books in two days: the exact step-by-step we use.",
        "price": 79_000, "file": "masagi-month-end-checklist.pdf", "icon": "✓",
    },
    "onboarding-course": {
        "name": "MASAGI HV Onboarding Mini-Course", "kind": "Course · 5 lessons",
        "tagline": "From empty database to first consolidated report, in five lessons.",
        "price": 299_000, "file": "masagi-onboarding-course.zip", "icon": "◮",
    },
}
BUNDLE = {
    "id": "founders-toolkit", "name": "Founder's Toolkit — everything",
    "tagline": "All five products. Every future update included.",
    "price": 499_000, "items": list(PRODUCTS.keys()),
}
MEMBERSHIP = {
    "id": "insider-membership", "name": "MASAGI Insider Membership",
    "tagline": "Everything in the toolkit + every new template and guide we release, "
               "while your membership is active.",
    "price": 99_000, "per": "month", "items": list(PRODUCTS.keys()),
}


# ---------------------------------------------------------------------------
# DB
# ---------------------------------------------------------------------------

def _db():
    os.makedirs(STORE_DIR, exist_ok=True)
    conn = sqlite3.connect(STORE_DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def _init_db():
    conn = _db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS customers(
      id INTEGER PRIMARY KEY, email TEXT UNIQUE NOT NULL, name TEXT DEFAULT '',
      password_hash TEXT NOT NULL, created_at TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS orders(
      id TEXT PRIMARY KEY, email TEXT NOT NULL, name TEXT DEFAULT '',
      customer_id INTEGER, plan TEXT NOT NULL, total INTEGER NOT NULL,
      status TEXT NOT NULL DEFAULT 'pending', method TEXT DEFAULT '',
      invoice_no TEXT DEFAULT '', created_at TEXT NOT NULL, paid_at TEXT);
    CREATE TABLE IF NOT EXISTS order_items(
      order_id TEXT NOT NULL, product_id TEXT NOT NULL, price INTEGER NOT NULL);
    CREATE TABLE IF NOT EXISTS entitlements(
      id INTEGER PRIMARY KEY, customer_id INTEGER NOT NULL, product_id TEXT NOT NULL,
      order_id TEXT NOT NULL, granted_at TEXT NOT NULL, expires_at TEXT,
      UNIQUE(customer_id, product_id, order_id));
    CREATE TABLE IF NOT EXISTS download_tokens(
      token_hash TEXT PRIMARY KEY, entitlement_id INTEGER NOT NULL,
      expires_at TEXT NOT NULL, uses INTEGER NOT NULL DEFAULT 0,
      max_uses INTEGER NOT NULL);
    """)
    conn.commit()
    conn.close()


def _now():
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")


# ---------------------------------------------------------------------------
# Private file seeding (real, working deliverables)
# ---------------------------------------------------------------------------

def _tiny_pdf(title, lines):
    """Dependency-free single-page PDF."""
    text = ["BT /F1 18 Tf 56 780 Td (%s) Tj ET" % title.replace("(", "[").replace(")", "]")]
    y = 744
    for ln in lines:
        text.append("BT /F1 11 Tf 56 %d Td (%s) Tj ET" % (y, ln.replace("(", "[").replace(")", "]")))
        y -= 20
    stream = "\n".join(text).encode("latin-1", "replace")
    objs = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] /Contents 4 0 R "
        b"/Resources << /Font << /F1 5 0 R >> >> >>",
        b"<< /Length %d >>\nstream\n%s\nendstream" % (len(stream), stream),
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    out, offsets = io.BytesIO(), []
    out.write(b"%PDF-1.4\n")
    for i, obj in enumerate(objs, 1):
        offsets.append(out.tell())
        out.write(b"%d 0 obj\n" % i + obj + b"\nendobj\n")
    xref = out.tell()
    out.write(b"xref\n0 %d\n0000000000 65535 f \n" % (len(objs) + 1))
    for off in offsets:
        out.write(b"%010d 00000 n \n" % off)
    out.write(b"trailer\n<< /Size %d /Root 1 0 R >>\nstartxref\n%d\n%%%%EOF" % (len(objs) + 1, xref))
    return out.getvalue()


def _seed_files():
    os.makedirs(FILES_DIR, exist_ok=True)
    os.makedirs(OUTBOX_DIR, exist_ok=True)

    def path(name):
        return os.path.join(FILES_DIR, name)

    if not os.path.exists(path("masagi-coa-template.xlsx")):
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "Chart of Accounts"
        ws.append(["Code", "Account name", "Type", "Notes"])
        for row in [("1100", "Cash & bank", "Asset", "One row per bank account"),
                    ("1200", "Accounts receivable", "Asset", ""),
                    ("2100", "Accounts payable", "Liability", ""),
                    ("3100", "Share capital", "Equity", ""),
                    ("4100", "Sales revenue", "Revenue", ""),
                    ("5100", "Cost of goods sold", "Expense", ""),
                    ("6100", "Salaries & wages", "Expense", ""),
                    ("6210", "Rent", "Expense", "")]:
            ws.append(list(row))
        wb.save(path("masagi-coa-template.xlsx"))

    if not os.path.exists(path("masagi-budget-planner.xlsx")):
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "Budget 2026"
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        ws.append(["Account"] + months + ["Total"])
        for acc in ["Sales revenue", "COGS", "Salaries", "Rent", "Marketing", "Other opex"]:
            ws.append([acc] + [0] * 12 + ["=SUM(B%d:M%d)" % (ws.max_row + 1, ws.max_row + 1)])
        wb.save(path("masagi-budget-planner.xlsx"))

    if not os.path.exists(path("masagi-sme-finance-playbook.pdf")):
        with open(path("masagi-sme-finance-playbook.pdf"), "wb") as f:
            f.write(_tiny_pdf("The SME Finance Playbook", [
                "MASAGI - Precise by default", "",
                "1. Know your numbers weekly, not quarterly.",
                "2. One chart of accounts, applied everywhere.",
                "3. Book per entity, report consolidated.",
                "4. The bank statement is the source of truth - import it, don't re-type it.",
                "5. Budgets are promises; track realization monthly.",
                "6. Close in two days: checklist beats memory.",
                "7. Let the system type, let people decide.", "",
                "Full edition ships with worked examples for trading, services and F&B.",
            ]))

    if not os.path.exists(path("masagi-month-end-checklist.pdf")):
        with open(path("masagi-month-end-checklist.pdf"), "wb") as f:
            f.write(_tiny_pdf("Month-End Close Checklist", [
                "Day 1, morning:  [ ] Import all bank statements   [ ] Book pending journals",
                "Day 1, afternoon:[ ] Reconcile AR / AP            [ ] Post accruals",
                "Day 2, morning:  [ ] Review anomalies flagged     [ ] Intercompany check",
                "Day 2, afternoon:[ ] Freeze the period            [ ] Export the board pack",
                "", "Tip: anything you did twice, automate before next month.",
            ]))

    if not os.path.exists(path("masagi-onboarding-course.zip")):
        lessons = {
            "lesson-1-setup.md": "# Lesson 1 — Setup\nCreate your database, apply the standard COA.",
            "lesson-2-first-journals.md": "# Lesson 2 — First journals\nDraft, review, post.",
            "lesson-3-bank-import.md": "# Lesson 3 — Bank import\nPaste mutasi, approve, done.",
            "lesson-4-budgets.md": "# Lesson 4 — Budgets\nPlan monthly, track realization.",
            "lesson-5-helicopter-view.md": "# Lesson 5 — Helicopter view\nConsolidate and decide.",
        }
        with zipfile.ZipFile(path("masagi-onboarding-course.zip"), "w") as z:
            for name, body in lessons.items():
                z.writestr("masagi-onboarding-course/" + name, body)


def init_app(app):
    _init_db()
    _seed_files()


# ---------------------------------------------------------------------------
# Helpers: signing, email, customer session
# ---------------------------------------------------------------------------

def _sign(payload: str) -> str:
    return hmac.new(DOKU_SECRET_KEY.encode(), payload.encode(), hashlib.sha256).hexdigest()


def _b64(s: bytes) -> str:
    return base64.urlsafe_b64encode(s).decode().rstrip("=")


def _unb64(s: str) -> bytes:
    return base64.urlsafe_b64decode(s + "=" * (-len(s) % 4))


def _make_token(kind, ref, hours):
    exp = int(time.time()) + hours * 3600
    payload = "%s|%s|%d" % (kind, ref, exp)
    return _b64(payload.encode()) + "." + _sign(payload)[:32]


def _read_token(token, kind):
    try:
        body, sig = token.split(".", 1)
        payload = _unb64(body).decode()
        if not hmac.compare_digest(_sign(payload)[:32], sig):
            return None
        k, ref, exp = payload.split("|")
        if k != kind or int(exp) < time.time():
            return None
        return ref
    except Exception:
        return None


def _send_email(to, subject, body):
    """Mock transactional email: writes to data/store_outbox/.
    # EMAIL: replace this body with Resend (https://api.resend.com/emails,
    # Authorization: Bearer RESEND_API_KEY) or Postmark / SMTP. Nothing else
    # in the pipeline needs to change."""
    fname = "%s_%s.eml.txt" % (datetime.utcnow().strftime("%Y%m%d-%H%M%S-%f"),
                               to.replace("@", "_at_"))
    with open(os.path.join(OUTBOX_DIR, fname), "w", encoding="utf-8") as f:
        f.write("To: %s\nSubject: %s\nDate: %s\n\n%s\n" % (to, subject, _now(), body))


def current_customer():
    cid = session.get("customer_id")
    if not cid:
        return None
    conn = _db()
    try:
        return conn.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
    finally:
        conn.close()


def _plan_details(plan, item_ids):
    """Resolve a checkout request into (items, total, label)."""
    if plan == "bundle":
        return BUNDLE["items"], BUNDLE["price"], BUNDLE["name"]
    if plan == "membership":
        return MEMBERSHIP["items"], MEMBERSHIP["price"], MEMBERSHIP["name"]
    items = [i for i in (item_ids or []) if i in PRODUCTS]
    return items, sum(PRODUCTS[i]["price"] for i in items), "One-time purchase"


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------

@bp.get("/store")
def store_page():
    return send_from_directory(STATIC_DIR, "store.html")


@bp.get("/store/pay/<order_id>")
def pay_page(order_id):
    return send_from_directory(STATIC_DIR, "store-checkout.html")


@bp.get("/store/thanks")
def thanks_page():
    return send_from_directory(STATIC_DIR, "store-thanks.html")


@bp.get("/downloads")
def downloads_page():
    return send_from_directory(STATIC_DIR, "store-downloads.html")


# ---------------------------------------------------------------------------
# Catalog + checkout
# ---------------------------------------------------------------------------

@bp.get("/api/store/products")
def api_products():
    return jsonify({
        "products": [dict(p, id=k) for k, p in PRODUCTS.items()],
        "bundle": BUNDLE, "membership": MEMBERSHIP,
    })


@bp.post("/api/store/checkout")
def api_checkout():
    """Step 2-3 of the pipeline: email capture -> pending order -> payment URL.
    # DOKU: in production this is where you call POST
    #   https://api.doku.com/checkout/v1/payment   (or api-sandbox.doku.com)
    # with headers Client-Id: DOKU_CLIENT_ID, Request-Id, Request-Timestamp and
    # Signature: HMACSHA256=... — and return response.payment.url instead of
    # the local sandbox page below."""
    d = request.get_json(force=True)
    email = (d.get("email") or "").strip().lower()
    if "@" not in email or "." not in email.split("@")[-1]:
        return jsonify({"error": "Please enter a valid email address."}), 400
    plan = d.get("plan") or "items"
    items, total, label = _plan_details(plan, d.get("items"))
    if not items:
        return jsonify({"error": "Nothing to buy — pick a product first."}), 400

    order_id = "MSG-" + secrets.token_hex(5).upper()
    conn = _db()
    try:
        conn.execute(
            "INSERT INTO orders(id,email,name,plan,total,status,created_at) VALUES(?,?,?,?,?,?,?)",
            (order_id, email, (d.get("name") or "").strip(), plan, total, "pending", _now()))
        for pid in items:
            conn.execute("INSERT INTO order_items(order_id,product_id,price) VALUES(?,?,?)",
                         (order_id, pid, PRODUCTS[pid]["price"]))
        conn.commit()
    finally:
        conn.close()
    return jsonify({"order_id": order_id, "payment_url": "/store/pay/" + order_id,
                    "label": label, "total": total})


@bp.get("/api/store/order/<order_id>")
def api_order(order_id):
    conn = _db()
    try:
        o = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        if not o:
            return jsonify({"error": "Order not found"}), 404
        items = conn.execute(
            "SELECT product_id, price FROM order_items WHERE order_id=?", (order_id,)).fetchall()
    finally:
        conn.close()
    return jsonify({
        "id": o["id"], "status": o["status"], "plan": o["plan"], "total": o["total"],
        "email": o["email"], "method": o["method"], "invoice_no": o["invoice_no"],
        "items": [{"id": i["product_id"], "name": PRODUCTS[i["product_id"]]["name"],
                   "price": i["price"]} for i in items if i["product_id"] in PRODUCTS],
    })


# ---------------------------------------------------------------------------
# Sandbox payment gateway (stands in for DOKU's hosted checkout page)
# ---------------------------------------------------------------------------

@bp.post("/api/store/gateway/pay")
def sandbox_gateway_pay():
    """SANDBOX ONLY — simulates the customer completing payment at DOKU.
    It builds the same signed notification a real gateway would send and runs
    it through the exact webhook processor. Remove in production; DOKU calls
    /api/store/webhook/doku itself."""
    d = request.get_json(force=True)
    order_id, method = d.get("order_id", ""), d.get("method", "QRIS")
    notification = json.dumps({
        "order": {"invoice_number": order_id},
        "transaction": {"status": "SUCCESS", "date": _now()},
        "channel": {"id": method},
    }, sort_keys=True)
    signature = "HMACSHA256=" + _sign(notification)
    ok, msg = _process_webhook(notification, signature)
    return (jsonify({"ok": True}) if ok else (jsonify({"error": msg}), 400))


@bp.post("/api/store/webhook/doku")
def doku_webhook():
    """Step 4: the webhook is the ONLY place access is granted.
    # DOKU: real notifications carry a Signature header computed from
    # Client-Id + Request-Id + Request-Timestamp + Request-Target + digest of
    # the body, HMAC-SHA256 with DOKU_SECRET_KEY. Verify exactly per
    # https://developers.doku.com/accept-payment/direct-api/signature ."""
    ok, msg = _process_webhook(request.get_data(as_text=True),
                               request.headers.get("Signature", ""))
    return (jsonify({"ok": True}) if ok else (jsonify({"error": msg}), 400))


def _process_webhook(raw_body, signature):
    expected = "HMACSHA256=" + _sign(raw_body)
    if not hmac.compare_digest(expected, signature or ""):
        return False, "Invalid signature"
    try:
        data = json.loads(raw_body)
        order_id = data["order"]["invoice_number"]
        status = data["transaction"]["status"]
        method = data.get("channel", {}).get("id", "")
    except (ValueError, KeyError):
        return False, "Malformed notification"
    if status != "SUCCESS":
        return False, "Ignored non-success status"

    conn = _db()
    try:
        o = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        if not o:
            return False, "Unknown order"
        if o["status"] == "paid":
            return True, "Already processed"  # idempotent

        # 1) auto-create the customer account (step 4 of the pipeline)
        cust = conn.execute("SELECT * FROM customers WHERE email=?", (o["email"],)).fetchone()
        new_password = None
        if not cust:
            new_password = secrets.token_urlsafe(9)
            conn.execute(
                "INSERT INTO customers(email,name,password_hash,created_at) VALUES(?,?,?,?)",
                (o["email"], o["name"], generate_password_hash(new_password), _now()))
            cust = conn.execute("SELECT * FROM customers WHERE email=?", (o["email"],)).fetchone()

        # 2) mark paid
        invoice_no = "INV-" + datetime.utcnow().strftime("%Y%m%d") + "-" + order_id[-4:]
        conn.execute("UPDATE orders SET status='paid', paid_at=?, method=?, customer_id=?, "
                     "invoice_no=? WHERE id=?",
                     (_now(), method, cust["id"], invoice_no, order_id))

        # 3) grant entitlements (membership entitlements expire; others don't)
        expires = ((datetime.utcnow() + timedelta(days=MEMBERSHIP_DAYS))
                   .strftime("%Y-%m-%d %H:%M:%S") if o["plan"] == "membership" else None)
        items = conn.execute("SELECT product_id FROM order_items WHERE order_id=?",
                             (order_id,)).fetchall()
        for it in items:
            conn.execute("INSERT OR IGNORE INTO entitlements(customer_id,product_id,order_id,"
                         "granted_at,expires_at) VALUES(?,?,?,?,?)",
                         (cust["id"], it["product_id"], order_id, _now(), expires))
        conn.commit()
    finally:
        conn.close()

    # 4) emails: receipt + magic login link (step 6)
    magic = _make_token("magic", str(cust["id"]), 48)
    login_url = request.url_root.rstrip("/") + "/api/store/magic?token=" + magic
    lines = ["Thank you for your purchase!", "",
             "Order:   %s" % order_id, "Invoice: %s" % invoice_no,
             "Total:   Rp %s" % format(o["total"], ",").replace(",", "."),
             "Method:  %s" % method, ""]
    if new_password:
        lines += ["We created your MASAGI account automatically:",
                  "  email:    %s" % o["email"],
                  "  password: %s   (change it after signing in)" % new_password, ""]
    lines += ["Open your downloads (magic link, valid 48h):", "  " + login_url]
    _send_email(o["email"], "Your MASAGI receipt — " + order_id, "\n".join(lines))
    return True, "ok"


# ---------------------------------------------------------------------------
# Customer auth (separate from ERP staff auth)
# ---------------------------------------------------------------------------

@bp.get("/api/store/magic")
def magic_login():
    cid = _read_token(request.args.get("token", ""), "magic")
    if not cid:
        return redirect("/downloads?expired=1")
    session["customer_id"] = int(cid)
    session.permanent = True
    return redirect("/downloads")


@bp.post("/api/store/login")
def customer_login():
    d = request.get_json(force=True)
    conn = _db()
    try:
        c = conn.execute("SELECT * FROM customers WHERE email=?",
                         ((d.get("email") or "").strip().lower(),)).fetchone()
    finally:
        conn.close()
    if not c or not check_password_hash(c["password_hash"], d.get("password") or ""):
        return jsonify({"error": "Wrong email or password."}), 401
    session["customer_id"] = c["id"]
    session.permanent = True
    return jsonify({"ok": True})


@bp.post("/api/store/magic-request")
def magic_request():
    email = (request.get_json(force=True).get("email") or "").strip().lower()
    conn = _db()
    try:
        c = conn.execute("SELECT * FROM customers WHERE email=?", (email,)).fetchone()
    finally:
        conn.close()
    if c:  # do not leak which emails exist
        url = request.url_root.rstrip("/") + "/api/store/magic?token=" + \
              _make_token("magic", str(c["id"]), 48)
        _send_email(email, "Your MASAGI sign-in link", "Sign in here (valid 48h):\n  " + url)
    return jsonify({"ok": True})


@bp.post("/api/store/logout")
def customer_logout():
    session.pop("customer_id", None)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# My Downloads (login-gated, per-user access control)
# ---------------------------------------------------------------------------

@bp.get("/api/store/me")
def api_me():
    c = current_customer()
    if not c:
        return jsonify({"error": "Not signed in"}), 401
    conn = _db()
    try:
        ents = conn.execute(
            "SELECT e.*, o.invoice_no FROM entitlements e JOIN orders o ON o.id=e.order_id "
            "WHERE e.customer_id=? ORDER BY e.granted_at DESC", (c["id"],)).fetchall()
        orders = conn.execute(
            "SELECT id,total,status,plan,invoice_no,created_at,paid_at FROM orders "
            "WHERE customer_id=? ORDER BY created_at DESC", (c["id"],)).fetchall()
    finally:
        conn.close()
    now = _now()
    seen, items = set(), []
    for e in ents:
        if e["product_id"] in seen or e["product_id"] not in PRODUCTS:
            continue
        seen.add(e["product_id"])
        p = PRODUCTS[e["product_id"]]
        active = (e["expires_at"] is None) or (e["expires_at"] > now)
        items.append({"entitlement_id": e["id"], "product_id": e["product_id"],
                      "name": p["name"], "kind": p["kind"], "icon": p["icon"],
                      "file": p["file"], "active": active, "expires_at": e["expires_at"],
                      "invoice_no": e["invoice_no"]})
    return jsonify({"email": c["email"], "name": c["name"], "products": items,
                    "orders": [dict(o) for o in orders]})


@bp.post("/api/store/download-link")
def make_download_link():
    """Signed, time-limited, use-limited link — the private file itself is
    never exposed. (Optional hardening from the brief: expiry + max uses.)"""
    c = current_customer()
    if not c:
        return jsonify({"error": "Not signed in"}), 401
    ent_id = request.get_json(force=True).get("entitlement_id")
    conn = _db()
    try:
        e = conn.execute("SELECT * FROM entitlements WHERE id=? AND customer_id=?",
                         (ent_id, c["id"])).fetchone()
        if not e:
            return jsonify({"error": "Not yours"}), 403
        if e["expires_at"] and e["expires_at"] < _now():
            return jsonify({"error": "Membership expired — renew to download."}), 403
        token = _make_token("dl", str(e["id"]), DOWNLOAD_TOKEN_HOURS)
        conn.execute("INSERT OR REPLACE INTO download_tokens(token_hash,entitlement_id,"
                     "expires_at,uses,max_uses) VALUES(?,?,?,0,?)",
                     (hashlib.sha256(token.encode()).hexdigest(), e["id"],
                      (datetime.utcnow() + timedelta(hours=DOWNLOAD_TOKEN_HOURS))
                      .strftime("%Y-%m-%d %H:%M:%S"), DOWNLOAD_TOKEN_MAX_USES))
        conn.commit()
    finally:
        conn.close()
    return jsonify({"url": "/api/store/download/" + token,
                    "expires_hours": DOWNLOAD_TOKEN_HOURS, "max_uses": DOWNLOAD_TOKEN_MAX_USES})


@bp.get("/api/store/download/<path:token>")
def signed_download(token):
    ent_id = _read_token(token, "dl")
    if not ent_id:
        return jsonify({"error": "Link expired — generate a fresh one from My Downloads."}), 410
    th = hashlib.sha256(token.encode()).hexdigest()
    conn = _db()
    try:
        t = conn.execute("SELECT * FROM download_tokens WHERE token_hash=?", (th,)).fetchone()
        if not t or t["uses"] >= t["max_uses"]:
            return jsonify({"error": "Link used up — generate a fresh one from My Downloads."}), 410
        e = conn.execute("SELECT * FROM entitlements WHERE id=?", (int(ent_id),)).fetchone()
        if not e or e["product_id"] not in PRODUCTS:
            return jsonify({"error": "Unknown product"}), 404
        conn.execute("UPDATE download_tokens SET uses=uses+1 WHERE token_hash=?", (th,))
        conn.commit()
        fname = PRODUCTS[e["product_id"]]["file"]
    finally:
        conn.close()
    return send_file(os.path.join(FILES_DIR, fname), as_attachment=True, download_name=fname)
