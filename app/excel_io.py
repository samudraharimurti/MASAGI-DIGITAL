"""MASAGI HV ERP - Excel import/export built on openpyxl."""
import io
from datetime import datetime, date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
BOLD = Font(bold=True)
NUM_FMT = "#,##0;[Red](#,##0)"
THIN = Border(bottom=Side(style="thin", color="CCCCCC"))


def _sheet(wb, title, report_title, subtitle=""):
    ws = wb.active if wb.active.max_row == 1 and wb.active.max_column == 1 else wb.create_sheet()
    ws.title = title[:31]
    ws["A1"] = report_title
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(color="666666", italic=True)
    return ws


def _header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _num(ws, row, col, value, bold=False):
    cell = ws.cell(row=row, column=col, value=round(value or 0, 2))
    cell.number_format = NUM_FMT
    if bold:
        cell.font = BOLD
    return cell


def _to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --------------------------------------------------------------------------
# Exports
# --------------------------------------------------------------------------

def export_trial_balance(tb, scope_label, period_label):
    wb = Workbook()
    ws = _sheet(wb, "Trial Balance", "Trial Balance — %s" % scope_label, period_label)
    _header_row(ws, 4, ["Code", "Account", "Type", "Debit", "Credit", "Balance"],
                [14, 40, 12, 18, 18, 18])
    r = 5
    # grouped by parent account when available: parent rows bold, children
    # indented; balances still reconcile because only leaf rows carry raw amounts
    display = tb.get("grouped") or [dict(x, level=0, is_group=False) for x in tb["rows"]]
    for row in display:
        is_group = row.get("is_group")
        indent = "    " * row.get("level", 0)
        c1 = ws.cell(row=r, column=1, value=row["code"])
        c2 = ws.cell(row=r, column=2, value=indent + row["name"])
        ws.cell(row=r, column=3, value=row["type"].title())
        _num(ws, r, 4, row["debit"], bold=is_group)
        _num(ws, r, 5, row["credit"], bold=is_group)
        _num(ws, r, 6, row["balance"], bold=is_group)
        if is_group:
            c1.font = BOLD
            c2.font = BOLD
        r += 1
    ws.cell(row=r, column=2, value="TOTAL").font = BOLD
    _num(ws, r, 4, tb["total_debit"], bold=True)
    _num(ws, r, 5, tb["total_credit"], bold=True)
    return _to_bytes(wb)


def export_receivables(aging, scope_label):
    wb = Workbook()
    ws = _sheet(wb, "Piutang", "AR Aging (Piutang) — %s" % scope_label,
                "As of %s" % aging["as_of"])
    _header_row(ws, 4, ["No", "Client", "Invoice", "Invoice Date", "Due Date",
                        "Amount", "Outstanding", "Not Due", "1-30 d", "31-60 d",
                        "61-90 d", "> 90 d", "Days Late", "Status"],
                [5, 28, 16, 13, 13, 16, 16, 14, 12, 12, 12, 12, 10, 20])
    bcol = {"not_due": 8, "d1_30": 9, "d31_60": 10, "d61_90": 11, "d90": 12}
    r = 5
    for i, it in enumerate(aging["items"], 1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=it["client"])
        ws.cell(row=r, column=3, value=it["invoice_no"])
        ws.cell(row=r, column=4, value=it["invoice_date"])
        ws.cell(row=r, column=5, value=it["due_date"])
        _num(ws, r, 6, it["amount"])
        _num(ws, r, 7, it["outstanding"])
        if it["bucket"]:
            _num(ws, r, bcol[it["bucket"]], it["outstanding"])
        ws.cell(row=r, column=13, value=it["days_overdue"] or 0)
        ws.cell(row=r, column=14, value=it["status_label"])
        r += 1
    ws.cell(row=r, column=2, value="TOTAL").font = BOLD
    _num(ws, r, 6, aging["total_amount"], bold=True)
    _num(ws, r, 7, aging["total_outstanding"], bold=True)
    for b, col in bcol.items():
        _num(ws, r, col, aging["buckets"][b], bold=True)
    r += 3
    ws.cell(row=r, column=2, value="AGING SUMMARY").font = BOLD
    r += 1
    for s in aging["summary"]:
        ws.cell(row=r, column=2, value=s["label"])
        _num(ws, r, 6, s["amount"])
        pc = ws.cell(row=r, column=7, value=(s["pct"] or 0) / 100.0)
        pc.number_format = "0.0%"
        r += 1
    ws.cell(row=r, column=2, value="TOTAL OUTSTANDING").font = BOLD
    _num(ws, r, 6, aging["total_outstanding"], bold=True)
    return _to_bytes(wb)


def export_payables(aging, scope_label):
    wb = Workbook()
    ws = _sheet(wb, "Hutang", "AP Aging (Hutang) — %s" % scope_label,
                "As of %s" % aging["as_of"])
    _header_row(ws, 4, ["No", "Vendor", "Bill", "Bill Date", "Due Date",
                        "Amount", "Outstanding", "Not Due", "1-30 d", "31-60 d",
                        "61-90 d", "> 90 d", "Days Late", "Status"],
                [5, 28, 16, 13, 13, 16, 16, 14, 12, 12, 12, 12, 10, 20])
    bcol = {"not_due": 8, "d1_30": 9, "d31_60": 10, "d61_90": 11, "d90": 12}
    r = 5
    for i, it in enumerate(aging["items"], 1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=it["vendor"])
        ws.cell(row=r, column=3, value=it["bill_no"])
        ws.cell(row=r, column=4, value=it["bill_date"])
        ws.cell(row=r, column=5, value=it["due_date"])
        _num(ws, r, 6, it["amount"])
        _num(ws, r, 7, it["outstanding"])
        if it["bucket"]:
            _num(ws, r, bcol[it["bucket"]], it["outstanding"])
        ws.cell(row=r, column=13, value=it["days_overdue"] or 0)
        ws.cell(row=r, column=14, value=it["status_label"])
        r += 1
    ws.cell(row=r, column=2, value="TOTAL").font = BOLD
    _num(ws, r, 6, aging["total_amount"], bold=True)
    _num(ws, r, 7, aging["total_outstanding"], bold=True)
    for b, col in bcol.items():
        _num(ws, r, col, aging["buckets"][b], bold=True)
    r += 3
    ws.cell(row=r, column=2, value="AGING SUMMARY").font = BOLD
    r += 1
    for s in aging["summary"]:
        ws.cell(row=r, column=2, value=s["label"])
        _num(ws, r, 6, s["amount"])
        pc = ws.cell(row=r, column=7, value=(s["pct"] or 0) / 100.0)
        pc.number_format = "0.0%"
        r += 1
    ws.cell(row=r, column=2, value="TOTAL OUTSTANDING").font = BOLD
    _num(ws, r, 6, aging["total_outstanding"], bold=True)
    return _to_bytes(wb)


def export_pnl(pnl, scope_label, period_label):
    wb = Workbook()
    ws = _sheet(wb, "Profit & Loss", "Profit & Loss — %s" % scope_label, period_label)
    _header_row(ws, 4, ["Code", "Account", "Amount"], [10, 42, 20])
    r = 5
    ws.cell(row=r, column=2, value="REVENUE").font = BOLD
    r += 1
    for row in pnl["revenue"]:
        ws.cell(row=r, column=1, value=row["code"])
        ws.cell(row=r, column=2, value=row["name"])
        _num(ws, r, 3, row["balance"])
        r += 1
    ws.cell(row=r, column=2, value="Total Revenue").font = BOLD
    _num(ws, r, 3, pnl["total_revenue"], bold=True)
    r += 2
    ws.cell(row=r, column=2, value="EXPENSES").font = BOLD
    r += 1
    for row in pnl["expense"]:
        ws.cell(row=r, column=1, value=row["code"])
        ws.cell(row=r, column=2, value=row["name"])
        _num(ws, r, 3, row["balance"])
        r += 1
    ws.cell(row=r, column=2, value="Total Expenses").font = BOLD
    _num(ws, r, 3, pnl["total_expense"], bold=True)
    r += 2
    ws.cell(row=r, column=2, value="NET PROFIT").font = Font(bold=True, size=12)
    _num(ws, r, 3, pnl["net_profit"], bold=True)
    return _to_bytes(wb)


def export_balance_sheet(bs, scope_label, period_label):
    wb = Workbook()
    ws = _sheet(wb, "Balance Sheet", "Balance Sheet — %s" % scope_label, period_label)
    _header_row(ws, 4, ["Code", "Account", "Amount"], [10, 42, 20])
    r = 5
    for section, rows, total in [
        ("ASSETS", bs["assets"], bs["total_assets"]),
        ("LIABILITIES", bs["liabilities"], bs["total_liabilities"]),
        ("EQUITY", bs["equity"], bs["total_equity"]),
    ]:
        ws.cell(row=r, column=2, value=section).font = BOLD
        r += 1
        for row in rows:
            ws.cell(row=r, column=1, value=row["code"])
            ws.cell(row=r, column=2, value=row["name"])
            _num(ws, r, 3, row["balance"])
            r += 1
        ws.cell(row=r, column=2, value="Total %s" % section.title()).font = BOLD
        _num(ws, r, 3, total, bold=True)
        r += 2
    return _to_bytes(wb)


def export_budget_vs_actual(bva, scope_label):
    wb = Workbook()
    ws = _sheet(wb, "Budget vs Realization",
                "Budget vs Realization — %s" % scope_label, "Year %s" % bva["year"])
    _header_row(ws, 4, ["Code", "Account", "Type", "Budget", "Realization", "Variance", "Used %"],
                [10, 36, 12, 18, 18, 18, 10])
    r = 5
    for row in bva["rows"]:
        ws.cell(row=r, column=1, value=row["code"])
        ws.cell(row=r, column=2, value=row["name"])
        ws.cell(row=r, column=3, value=row["type"].title())
        _num(ws, r, 4, row["budget"])
        _num(ws, r, 5, row["actual"])
        _num(ws, r, 6, row["variance"])
        if row["used_pct"] is not None:
            ws.cell(row=r, column=7, value=row["used_pct"] / 100).number_format = "0.0%"
        r += 1
    return _to_bytes(wb)


def export_budget_grid(rows, scope_label, year):
    """rows: [{code, name, amounts:[12]}]"""
    wb = Workbook()
    ws = _sheet(wb, "Budget %s" % year, "Budget — %s" % scope_label, "Year %s" % year)
    _header_row(ws, 4, ["Account Code", "Account Name"] + MONTHS + ["Total"],
                [14, 32] + [13] * 13)
    r = 5
    for row in rows:
        ws.cell(row=r, column=1, value=row["code"])
        ws.cell(row=r, column=2, value=row["name"])
        for m in range(12):
            _num(ws, r, 3 + m, row["amounts"][m])
        _num(ws, r, 15, sum(row["amounts"]), bold=True)
        r += 1
    return _to_bytes(wb)


def export_journals(entries, scope_label, period_label):
    """entries: [{entry_no, date, status, description, reference, lines:[{account_code, account_name, project_code, description, debit, credit}]}]"""
    wb = Workbook()
    ws = _sheet(wb, "Journal Entries", "Journal Entries — %s" % scope_label, period_label)
    _header_row(ws, 4, ["Entry No", "Date", "Status", "Entry Description", "Account Code",
                        "Account Name", "Project", "Line Description", "Debit", "Credit"],
                [16, 12, 9, 32, 12, 28, 12, 30, 17, 17])
    r = 5
    for e in entries:
        for ln in e["lines"]:
            ws.cell(row=r, column=1, value=e["entry_no"])
            ws.cell(row=r, column=2, value=e["date"])
            ws.cell(row=r, column=3, value=e["status"])
            ws.cell(row=r, column=4, value=e["description"])
            ws.cell(row=r, column=5, value=ln["account_code"])
            ws.cell(row=r, column=6, value=ln["account_name"])
            ws.cell(row=r, column=7, value=ln.get("project_code") or "")
            ws.cell(row=r, column=8, value=ln["description"])
            _num(ws, r, 9, ln["debit"])
            _num(ws, r, 10, ln["credit"])
            r += 1
    return _to_bytes(wb)


def export_coa(accounts, scope_label):
    wb = Workbook()
    ws = _sheet(wb, "Chart of Accounts", "Chart of Accounts — %s" % scope_label)
    _header_row(ws, 4, ["Code", "Name", "Type", "Parent Code", "Intercompany", "Active"],
                [10, 40, 12, 12, 13, 8])
    r = 5
    for a in accounts:
        ws.cell(row=r, column=1, value=a["code"])
        ws.cell(row=r, column=2, value=a["name"])
        ws.cell(row=r, column=3, value=a["type"].title())
        ws.cell(row=r, column=4, value=a["parent_code"] or "")
        ws.cell(row=r, column=5, value="Y" if a["is_intercompany"] else "")
        ws.cell(row=r, column=6, value="Y" if a["is_active"] else "N")
        r += 1
    return _to_bytes(wb)


def export_project_performance(rows, scope_label, year):
    wb = Workbook()
    ws = _sheet(wb, "Project Performance",
                "Project Performance — %s" % scope_label, "Year %s" % year)
    _header_row(ws, 4, ["Company", "Code", "Project", "Status", "Revenue", "Expense",
                        "Profit", "Margin %", "Budget Rev", "Budget Exp"],
                [10, 12, 32, 10, 18, 18, 18, 10, 18, 18])
    r = 5
    for p in rows:
        ws.cell(row=r, column=1, value=p["company"])
        ws.cell(row=r, column=2, value=p["code"])
        ws.cell(row=r, column=3, value=p["name"])
        ws.cell(row=r, column=4, value=p["status"])
        _num(ws, r, 5, p["revenue"])
        _num(ws, r, 6, p["expense"])
        _num(ws, r, 7, p["profit"])
        ws.cell(row=r, column=8, value=p["margin_pct"] / 100).number_format = "0.0%"
        _num(ws, r, 9, p["budget_revenue"])
        _num(ws, r, 10, p["budget_expense"])
        r += 1
    return _to_bytes(wb)


def export_cash_flow(cf, scope_label):
    wb = Workbook()
    ws = _sheet(wb, "Cash Flow", "Cash Flow Analysis — %s" % scope_label, "Year %s" % cf["year"])
    _header_row(ws, 4, ["Month", "Cash In", "Cash Out", "Net", "Ending Balance"],
                [12, 20, 20, 20, 20])
    ws.cell(row=5, column=1, value="Opening").font = BOLD
    _num(ws, 5, 5, cf["opening_balance"], bold=True)
    r = 6
    for m in cf["monthly"]:
        ws.cell(row=r, column=1, value=MONTHS[m["month"] - 1])
        _num(ws, r, 2, m["cash_in"])
        _num(ws, r, 3, m["cash_out"])
        _num(ws, r, 4, m["net"])
        _num(ws, r, 5, m["ending"])
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = BOLD
    _num(ws, r, 2, cf["total_in"], bold=True)
    _num(ws, r, 3, cf["total_out"], bold=True)
    _num(ws, r, 4, cf["net_change"], bold=True)
    _num(ws, r, 5, cf["closing_balance"], bold=True)
    r += 2
    ws.cell(row=r, column=1, value="SOURCES OF CASH").font = BOLD
    ws.cell(row=r, column=4, value="USES OF CASH").font = BOLD
    r += 1
    for i in range(max(len(cf["sources"]), len(cf["uses"]))):
        if i < len(cf["sources"]):
            s = cf["sources"][i]
            ws.cell(row=r + i, column=1, value="%s %s" % (s["code"], s["name"]))
            _num(ws, r + i, 2, s["amount"])
        if i < len(cf["uses"]):
            u = cf["uses"][i]
            ws.cell(row=r + i, column=4, value="%s %s" % (u["code"], u["name"]))
            _num(ws, r + i, 5, u["amount"])
    return _to_bytes(wb)


# --------------------------------------------------------------------------
# Templates
# --------------------------------------------------------------------------

def template_journals():
    wb = Workbook()
    ws = _sheet(wb, "Journals", "Journal Import Template",
                "One row per line. Rows sharing the same Entry Ref form one entry; debits must equal credits.")
    _header_row(ws, 4, ["Entry Ref", "Date (YYYY-MM-DD)", "Description", "Account Code",
                        "Line Description", "Debit", "Credit", "Project Code"],
                [12, 18, 32, 14, 30, 16, 16, 14])
    sample = [
        ("INV-001", "2026-06-01", "Customer invoice", "1200", "Invoice #123", 50000000, 0, "PRJ-APP"),
        ("INV-001", "2026-06-01", "Customer invoice", "4100", "Invoice #123", 0, 50000000, "PRJ-APP"),
        ("PAY-001", "2026-06-05", "Office supplies", "6600", "Stationery", 1500000, 0, ""),
        ("PAY-001", "2026-06-05", "Office supplies", "1120", "Bank payment", 0, 1500000, ""),
    ]
    for i, row in enumerate(sample):
        for j, v in enumerate(row, start=1):
            ws.cell(row=5 + i, column=j, value=v)
    return _to_bytes(wb)


def template_coa():
    wb = Workbook()
    ws = _sheet(wb, "COA", "Chart of Accounts Import Template",
                "Type must be one of: Asset, Liability, Equity, Revenue, Expense")
    _header_row(ws, 4, ["Code", "Name", "Type", "Parent Code", "Intercompany (Y/N)"],
                [10, 40, 12, 12, 16])
    ws.append([])
    for row in [("1130", "Petty Cash Branch", "Asset", "1100", ""),
                ("6800", "Travel & Entertainment", "Expense", "6000", "")]:
        ws.append(row)
    return _to_bytes(wb)


def template_budget(year):
    wb = Workbook()
    ws = _sheet(wb, "Budget", "Budget Import Template",
                "Amounts per month for year %s. Account Code must exist in the company COA." % year)
    _header_row(ws, 4, ["Account Code", "Project Code (optional)"] + MONTHS, [14, 20] + [13] * 12)
    ws.cell(row=5, column=1, value="6100")
    for m in range(12):
        ws.cell(row=5, column=3 + m, value=500000000)
    return _to_bytes(wb)


# --------------------------------------------------------------------------
# Imports
# --------------------------------------------------------------------------

def upsert_budget(conn, company_id, account_id, project_id, year, month, amount):
    """NULL-safe budget upsert (UNIQUE constraint treats NULL project_id rows as distinct)."""
    cur = conn.execute(
        "UPDATE budgets SET amount=? WHERE company_id=? AND account_id=?"
        " AND project_id IS ? AND year=? AND month=?",
        (round(amount, 2), company_id, account_id, project_id, year, month),
    )
    if cur.rowcount == 0:
        conn.execute(
            "INSERT INTO budgets (company_id, account_id, project_id, year, month, amount)"
            " VALUES (?,?,?,?,?,?)",
            (company_id, account_id, project_id, year, month, round(amount, 2)),
        )


def _cell_str(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _cell_num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return 0.0


def _find_header(ws, first_col_name):
    for row in ws.iter_rows(min_row=1, max_row=10):
        if _cell_str(row[0].value).lower().startswith(first_col_name.lower()):
            return row[0].row
    return None


def import_journals(conn, company_id, stream, created_by):
    """Returns (created_entries, errors)."""
    wb = load_workbook(stream, data_only=True)
    ws = wb.active
    hdr = _find_header(ws, "Entry Ref")
    if not hdr:
        return 0, ["Header row not found — first column must be 'Entry Ref'. Use the template."]
    accounts = {r["code"]: r["id"] for r in conn.execute(
        "SELECT id, code FROM accounts WHERE company_id=? AND is_active=1", (company_id,))}
    projects = {r["code"]: r["id"] for r in conn.execute(
        "SELECT id, code FROM projects WHERE company_id=?", (company_id,))}

    groups, errors = {}, []
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        if not row or not _cell_str(row[0]):
            continue
        ref = _cell_str(row[0])
        entry_date = _cell_str(row[1])
        try:
            datetime.strptime(entry_date, "%Y-%m-%d")
        except ValueError:
            errors.append("Entry %s: invalid date '%s' (use YYYY-MM-DD)" % (ref, entry_date))
            continue
        code = _cell_str(row[3])
        if code not in accounts:
            errors.append("Entry %s: account code '%s' not found in this company" % (ref, code))
            continue
        pcode = _cell_str(row[7]) if len(row) > 7 else ""
        if pcode and pcode not in projects:
            errors.append("Entry %s: project code '%s' not found" % (ref, pcode))
            continue
        g = groups.setdefault(ref, {"date": entry_date, "description": _cell_str(row[2]), "lines": []})
        g["lines"].append({
            "account_id": accounts[code],
            "description": _cell_str(row[4]),
            "debit": _cell_num(row[5]),
            "credit": _cell_num(row[6]),
            "project_id": projects.get(pcode),
        })

    if errors:
        return 0, errors

    created = 0
    for ref, g in groups.items():
        total_d = round(sum(l["debit"] for l in g["lines"]), 2)
        total_c = round(sum(l["credit"] for l in g["lines"]), 2)
        if abs(total_d - total_c) > 0.01:
            errors.append("Entry %s is not balanced (debit %.2f vs credit %.2f) — skipped" % (ref, total_d, total_c))
            continue
        if total_d == 0:
            errors.append("Entry %s has zero amount — skipped" % ref)
            continue
        n = conn.execute(
            "SELECT COUNT(*)+1 FROM journal_entries WHERE company_id=?", (company_id,)
        ).fetchone()[0]
        entry_no = "IMP-%s-%05d" % (g["date"][:7].replace("-", ""), n)
        cur = conn.execute(
            "INSERT INTO journal_entries (company_id, entry_no, date, description, reference, status, source, created_by)"
            " VALUES (?,?,?,?,?,'draft','excel',?)",
            (company_id, entry_no, g["date"], g["description"], ref, created_by),
        )
        for l in g["lines"]:
            conn.execute(
                "INSERT INTO journal_lines (entry_id, account_id, project_id, description, debit, credit)"
                " VALUES (?,?,?,?,?,?)",
                (cur.lastrowid, l["account_id"], l["project_id"], l["description"],
                 round(l["debit"], 2), round(l["credit"], 2)),
            )
        created += 1
    conn.commit()
    return created, errors


def import_coa(conn, company_id, stream):
    wb = load_workbook(stream, data_only=True)
    ws = wb.active
    hdr = _find_header(ws, "Code")
    if not hdr:
        return 0, 0, ["Header row not found — first column must be 'Code'. Use the template."]
    valid_types = {"asset", "liability", "equity", "revenue", "expense"}
    created = updated = 0
    errors = []
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        if not row or not _cell_str(row[0]):
            continue
        code, name = _cell_str(row[0]), _cell_str(row[1])
        typ = _cell_str(row[2]).lower()
        parent = _cell_str(row[3]) if len(row) > 3 else ""
        if not parent and "-" in code:
            parent = code.rsplit("-", 1)[0]  # derivative: 5100-01-01 -> 5100-01
        ic = 1 if (len(row) > 4 and _cell_str(row[4]).upper() in ("Y", "YES", "1")) else 0
        if typ not in valid_types:
            errors.append("Account %s: invalid type '%s'" % (code, typ))
            continue
        if code.count("-") > 2:
            errors.append("Account %s: maximum 3 levels (e.g. 5100-01-01)" % code)
            continue
        existing = conn.execute(
            "SELECT id FROM accounts WHERE company_id=? AND code=?", (company_id, code)
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE accounts SET name=?, type=?, parent_code=?, is_intercompany=? WHERE id=?",
                (name, typ, parent or None, ic, existing["id"]))
            updated += 1
        else:
            conn.execute(
                "INSERT INTO accounts (company_id, code, name, type, parent_code, is_intercompany)"
                " VALUES (?,?,?,?,?,?)",
                (company_id, code, name, typ, parent or None, ic))
            created += 1
    conn.commit()
    return created, updated, errors


def import_budget(conn, company_id, year, stream):
    wb = load_workbook(stream, data_only=True)
    ws = wb.active
    hdr = _find_header(ws, "Account Code")
    if not hdr:
        return 0, ["Header row not found — first column must be 'Account Code'. Use the template."]
    accounts = {r["code"]: r["id"] for r in conn.execute(
        "SELECT id, code FROM accounts WHERE company_id=?", (company_id,))}
    projects = {r["code"]: r["id"] for r in conn.execute(
        "SELECT id, code FROM projects WHERE company_id=?", (company_id,))}
    saved, errors = 0, []
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        if not row or not _cell_str(row[0]):
            continue
        code = _cell_str(row[0])
        if code not in accounts:
            errors.append("Account code '%s' not found — row skipped" % code)
            continue
        pcode = _cell_str(row[1]) if len(row) > 1 else ""
        if pcode and pcode not in projects:
            errors.append("Project code '%s' not found — row skipped" % pcode)
            continue
        pid = projects.get(pcode)
        for m in range(12):
            amount = _cell_num(row[2 + m]) if len(row) > 2 + m else 0.0
            upsert_budget(conn, company_id, accounts[code], pid, year, m + 1, amount)
        saved += 1
    conn.commit()
    return saved, errors
